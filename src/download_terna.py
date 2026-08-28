#!/usr/bin/env python3
"""Download solar data from Terna and save it as CSV.

Choose one of these commands:

- "actual-generation" downloads the solar electricity produced in Italy.
- "installed-renewables" downloads monthly changes in installed capacity.
- "all" downloads both.

All commands use --start and --end. Each value can be a complete date or only
a year:

- 2024-01-20 means that exact date;
- 2024 means January 1 when used with --start;
- 2024 means December 31 when used with --end.

Actual generation uses the resulting complete dates. Installed renewables
uses only their years because its files are annual exports. The all command
uses the complete dates for actual generation and their years for the annual
exports.

Actual generation requires a Key and Secret. Create them by registering an
application at https://developer.terna.it/, then store them in
TERNA_CLIENT_ID and TERNA_CLIENT_SECRET. Installed-renewables files do
not require credentials.

Files are saved in share/terna unless --output-dir is provided.

Examples:

    python src/download_terna.py actual-generation --start 2024-01-20 --end 2026-06-30

    python src/download_terna.py installed-renewables --start 2024 --end 2026

    python src/download_terna.py all --start 2024-01-20 --end 2026-06-30

    python src/download_terna.py actual-generation \
        --start 2024-01-20 \
        --end 2026-06-30 \
        --output-dir path/to/output-dir
"""

from __future__ import annotations

import argparse
import csv
import os
import re
import sys
import time
from dataclasses import dataclass
from datetime import date, datetime, timedelta, timezone
from pathlib import Path
from typing import Any

import requests
from openpyxl import load_workbook
from requests.adapters import HTTPAdapter
from urllib3.util.retry import Retry


TOKEN_URL = "https://api.terna.it/public-api/access-token"
ACTUAL_GENERATION_URL = (
    "https://api.terna.it/generation/v2.0/actual-generation"
)
DOWNLOAD_CENTER_RECORDS_URL = (
    "https://dati.terna.it/api/sitecore/dati/downloadcenter/records"
)
SOURCE = "Photovoltaic"
CLIENT_ID_ENV = "TERNA_CLIENT_ID"
CLIENT_SECRET_ENV = "TERNA_CLIENT_SECRET"
DEFAULT_REQUEST_INTERVAL = 2.0
QPS_RETRIES = 6
MAX_DOWNLOAD_CENTER_ROWS = 1_048_573
FIRST_AVAILABLE_RENEWABLE_YEAR = 2024
INSTALLED_RENEWABLE_DATASET = "RenewableSources"
INSTALLED_RENEWABLE_TREND = "Totale Consuntivi"

PROJECT_ROOT = Path(__file__).resolve().parents[1]
DEFAULT_OUTPUT_DIR = PROJECT_ROOT / "share" / "terna"
ACTUAL_GENERATION_FILENAME = "actual_generation.csv"
INSTALLED_RENEWABLES_FILENAME = "installed_renewables_{year}.csv"

INSTALLED_RENEWABLE_SOURCE_HEADERS = (
    "Anno",
    "Mese",
    "Zona mercato",
    "Regione",
    "Provincia",
    "Classe di potenza dell'impianto",
    "Tipo Variazione",
    "Fonte",
    "Livello tensione",
    "Potenza attiva nominale (MW)",
    "Numero impianti"
)
INSTALLED_RENEWABLE_OUTPUT_HEADERS = (
    "year",
    "month",
    "market_zone",
    "region",
    "province",
    "plant_power_class",
    "change_type",
    "source",
    "voltage_level",
    "nominal_active_power_mw",
    "plant_count"
)
OUTPUT_COLUMNS = (
    "time_utc",
    "time_local",
    "utc_offset",
    "timezone",
    "primary_source",
    "actual_generation_gwh"
)


@dataclass
class TernaClient:
    """Connect to Terna and respect its request limits."""

    client_id: str
    client_secret: str
    timeout: float
    request_interval: float

    def __post_init__(self) -> None:
        """Prepare the connection and login information."""

        self.session = build_session()
        self.access_token: str | None = None
        self.expires_at = 0.0
        self.last_request_at: float | None = None

    def wait_for_request_slot(self) -> None:
        """Wait if the previous request was sent too recently."""

        if self.last_request_at is not None:
            elapsed = time.monotonic() - self.last_request_at
            remaining = self.request_interval - elapsed
            if remaining > 0.0:
                time.sleep(remaining)
        self.last_request_at = time.monotonic()

    @staticmethod
    def is_qps_limit(response: requests.Response) -> bool:
        """Check whether Terna rejected a request sent too quickly."""

        return response.status_code == 429 or (
            response.status_code == 403
            and "over qps" in response.text.casefold()
        )

    def send(self, method: str, url: str, **kwargs: Any) -> requests.Response:
        """Send a request and retry when Terna reports too many requests."""

        for attempt in range(QPS_RETRIES):
            self.wait_for_request_slot()
            response = self.session.request(method, url, **kwargs)

            if not self.is_qps_limit(response):
                return response

            if attempt == QPS_RETRIES - 1:
                return response

            retry_after_text = response.headers.get("Retry-After", "")
            try:
                retry_after = float(retry_after_text)
            except ValueError:
                retry_after = 0.0

            backoff = min(60.0, 2.0 ** (attempt + 1))
            delay = max(self.request_interval, retry_after, backoff)
            print(
                "Terna QPS limit reached; retrying in "
                f"{delay:g} seconds ({attempt + 1}/{QPS_RETRIES - 1}).",
                file=sys.stderr,
                flush=True
            )
            time.sleep(delay)
        raise AssertionError("Unreachable code.")

    def refresh_token(self) -> None:
        """Get a new login token from Terna."""

        response = self.send(
            "POST",
            TOKEN_URL,
            data={
                "grant_type": "client_credentials",
                "client_id": self.client_id,
                "client_secret": self.client_secret
            },
            headers={"Content-Type": "application/x-www-form-urlencoded"},
            timeout=self.timeout
        )

        if not response.ok:
            raise RuntimeError(
                f"Terna token request failed with HTTP {response.status_code}: "
                f"{response.text[:500]}."
            )

        payload = response.json()
        token = payload.get("access_token")

        if not token:
            raise RuntimeError("Terna token response has no access_token.")

        lifetime = float(payload.get("expires_in", 300))
        self.access_token = str(token)
        self.expires_at = time.monotonic() + max(30.0, lifetime - 30.0)

    def token(self) -> str:
        """Return a valid login token and renew it when needed."""

        if self.access_token is None or time.monotonic() >= self.expires_at:
            self.refresh_token()
        assert self.access_token is not None, "Terna access token is missing."

        return self.access_token

    def actual_generation(self, start: date, end: date) -> list[dict[str, Any]]:
        """Get energy generation between two dates."""

        params = [
            ("dateFrom", start.strftime("%d/%m/%Y")),
            ("dateTo", end.strftime("%d/%m/%Y")),
            ("type", SOURCE)
        ]

        for attempt in range(2):
            response = self.send(
                "GET",
                ACTUAL_GENERATION_URL,
                params=params,
                headers={"Authorization": f"Bearer {self.token()}"},
                timeout=self.timeout
            )

            if response.status_code == 401 and attempt == 0:
                self.access_token = None
                continue

            if not response.ok:
                raise RuntimeError(
                    "Terna data request failed with HTTP "
                    f"{response.status_code}: {response.text[:800]}."
                )

            payload = response.json()
            status = str(payload.get("result", {}).get("status", ""))

            if status.casefold() != "completed":
                message = payload.get("result", {}).get("message", "unknown error")
                raise RuntimeError(f"Terna request was not completed: {message}.")

            rows = payload.get("actual_generation")

            if not isinstance(rows, list):
                raise RuntimeError("Terna response has no actual_generation list.")

            return rows

        raise RuntimeError("Terna authorization failed after refreshing the token.")


def add_date_arguments(parser: argparse.ArgumentParser) -> None:
    """Add --start and --end, which accept a date or a year."""

    parser.add_argument(
        "--start",
        required=True,
        help=(
            "First date of the requested period. Use YYYY-MM-DD for an exact "
            "date, or YYYY to start from January 1 of that year."
        )
    )
    parser.add_argument(
        "--end",
        required=True,
        help=(
            "Last date of the requested period. Use YYYY-MM-DD for an exact "
            "date, or YYYY to end on December 31 of that year."
        )
    )


def add_http_arguments(parser: argparse.ArgumentParser) -> None:
    """Add --timeout to set how long one web request may take."""

    parser.add_argument(
        "--timeout",
        type=float,
        default=120.0,
        help=(
            "Stop a web request if it takes more than this many seconds. "
            "Default: 120."
        )
    )


def add_output_directory_argument(parser: argparse.ArgumentParser) -> None:
    """Add --output-dir to define the folder where CSV files will be saved."""

    parser.add_argument(
        "--output-dir",
        type=Path,
        default=DEFAULT_OUTPUT_DIR,
        help="Folder where CSV files will be saved. Default: share/terna."
    )


def add_request_interval_argument(parser: argparse.ArgumentParser) -> None:
    """Add --request-interval to set the wait between Terna requests."""

    parser.add_argument(
        "--request-interval",
        "--pause",
        dest="request_interval",
        type=float,
        default=DEFAULT_REQUEST_INTERVAL,
        help=(
            "Seconds to wait between Terna requests. Increase this value if "
            "Terna reports too many requests. "
            f"Default: {DEFAULT_REQUEST_INTERVAL:g}."
        )
    )


def add_common_download_arguments(parser: argparse.ArgumentParser) -> None:
    """Add the period, output folder, and timeout used by every command."""

    add_date_arguments(parser)
    add_output_directory_argument(parser)
    add_http_arguments(parser)


def parse_args(argv: list[str] | None = None) -> argparse.Namespace:
    """Define the available commands and read the user's choices."""

    parser = argparse.ArgumentParser(
        description=__doc__,
        formatter_class=argparse.RawDescriptionHelpFormatter
    )
    commands = parser.add_subparsers(dest="command", required=True)

    actual = commands.add_parser(
        "actual-generation",
        help="Download the measured solar electricity produced in Italy.",
        description=(
            "Download actual solar generation for the period selected by "
            "--start and --end. Use a year alone to select the whole year."
        )
    )
    add_common_download_arguments(actual)
    add_request_interval_argument(actual)

    installed = commands.add_parser(
        "installed-renewables",
        help=(
            "Download yearly files with monthly changes in installed "
            "renewable capacity."
        ),
        description=(
            "Download installed-renewables files for the years selected by "
            "--start and --end. Days and months are not used in annual files."
        )
    )
    add_common_download_arguments(installed)

    both = commands.add_parser(
        "all",
        help="Download both actual generation and installed-renewables files.",
        description=(
            "Download actual generation for the selected period and the "
            "installed-renewables export for each year in that period."
        )
    )
    add_common_download_arguments(both)
    add_request_interval_argument(both)

    return parser.parse_args(argv)


def parse_date_or_year(
    value: str,
    option: str,
    *,
    end_of_year: bool
) -> date:
    """Convert YYYY-MM-DD to a date, or YYYY to the first or last day."""

    text = value.strip()

    try:
        if re.fullmatch(r"\d{4}", text):
            year = int(text)
            if end_of_year:
                return date(year, 12, 31)
            return date(year, 1, 1)

        if not re.fullmatch(r"\d{4}-\d{2}-\d{2}", text):
            raise ValueError("Unsupported date format.")

        return date.fromisoformat(text)
    except ValueError as exc:
        raise ValueError(
            f"{option} must use YYYY-MM-DD or YYYY: {value!r}."
        ) from exc


def validate_dates(start_text: str, end_text: str) -> tuple[date, date]:
    """Convert --start and --end to dates and check their order."""

    start = parse_date_or_year(
        start_text,
        "--start",
        end_of_year=False
    )
    end = parse_date_or_year(
        end_text,
        "--end",
        end_of_year=True
    )

    if end < start:
        raise ValueError("--end must not be earlier than --start.")

    return start, end


def create_month_list(start: date, end: date) -> list[tuple[date, date]]:
    """Split a date range into one part for each month."""

    ranges = []
    current = start

    while current <= end:
        if current.month == 12:
            next_month = date(current.year + 1, 1, 1)
        else:
            next_month = date(current.year, current.month + 1, 1)
        chunk_end = min(end, next_month - timedelta(days=1))
        ranges.append((current, chunk_end))
        current = chunk_end + timedelta(days=1)

    return ranges


def build_session(*, download_center: bool = False) -> requests.Session:
    """Create the connection used to call Terna."""

    retry = Retry(
        total=6,
        connect=6,
        read=6,
        status=6,
        backoff_factor=1.0,
        status_forcelist=(429, 500, 502, 503, 504),
        allowed_methods=frozenset({"GET", "POST"}),
        raise_on_status=False
    )
    adapter = HTTPAdapter(max_retries=retry)
    session = requests.Session()
    session.mount("https://", adapter)
    session.headers.update({"Accept": "application/json"})

    if not download_center:
        session.headers.update(
            {"User-Agent": "day-ahead-solar-forecasting/2.0"}
        )

    return session


def offset_from_text(value: str) -> timezone:
    """Convert text such as +01:00 into a UTC offset."""

    match = re.fullmatch(r"([+-])(\d{2}):(\d{2})", value.strip())

    if match is None:
        raise ValueError(f"Invalid UTC offset returned by Terna: {value!r}.")

    sign = 1 if match.group(1) == "+" else -1
    delta = timedelta(
        hours=sign * int(match.group(2)),
        minutes=sign * int(match.group(3))
    )

    return timezone(delta)


def parse_terna_timestamp(local_value: str, offset_value: str) -> datetime:
    """Convert a Terna date and time to UTC."""

    local = datetime.fromisoformat(local_value.strip().replace(" ", "T"))

    if local.tzinfo is None:
        local = local.replace(tzinfo=offset_from_text(offset_value))

    return local.astimezone(timezone.utc)


def parse_number(value: Any, field: str) -> float:
    """Convert a Terna value to a number and reject invalid values."""

    if value is None:
        raise ValueError(f"Terna returned null for {field}.")

    text = str(value).strip().replace(",", "")

    try:
        number = float(text)
    except ValueError as exc:
        raise ValueError(f"Invalid numeric value for {field}: {value!r}.") from exc

    if number < 0.0:
        raise ValueError(f"Negative value for {field}: {number}.")

    return number


def utc_text(value: datetime) -> str:
    """Convert a date and time to UTC text."""

    return value.astimezone(timezone.utc).isoformat(timespec="seconds").replace(
        "+00:00",
        "Z"
    )


def normalize_item(item: dict[str, Any]) -> dict[str, str]:
    """Convert one generation record to the CSV format."""

    required = {
        "date",
        "date_tz",
        "date_offset",
        "actual_generation_GWh",
        "primary_source"
    }
    missing = required - set(item)

    if missing:
        raise RuntimeError(f"Terna row is missing fields: {sorted(missing)}.")

    source = str(item["primary_source"]).strip()

    if source.casefold() != SOURCE.casefold():
        raise RuntimeError(f"Unexpected primary source returned by Terna: {source}.")

    local_text = str(item["date"]).strip().replace("T", " ")
    offset_text = str(item["date_offset"]).strip()
    instant_utc = parse_terna_timestamp(local_text, offset_text)
    generation = parse_number(
        item["actual_generation_GWh"],
        "actual_generation_GWh"
    )

    return {
        "time_utc": utc_text(instant_utc),
        "time_local": local_text,
        "utc_offset": offset_text,
        "timezone": str(item["date_tz"]).strip(),
        "primary_source": source,
        "actual_generation_gwh": format(generation, ".12g")
    }


def validate_and_sort(rows: list[dict[str, str]]) -> list[dict[str, str]]:
    """Check generation rows, remove duplicates, and sort them by time."""

    by_key: dict[tuple[str, str], dict[str, str]] = {}

    for row in rows:
        key = (row["time_utc"], row["primary_source"])
        previous = by_key.get(key)

        if previous is not None and previous != row:
            raise RuntimeError(f"Conflicting Terna rows for {key}.")
        by_key[key] = row

    result = sorted(
        by_key.values(),
        key=lambda row: (row["time_utc"], row["primary_source"])
    )

    if not result:
        raise RuntimeError("Terna returned no photovoltaic generation rows.")

    return result


def write_csv_atomic(rows: list[dict[str, str]], output: Path) -> None:
    """Save actual-generation rows to a CSV file."""

    output.parent.mkdir(parents=True, exist_ok=True)
    temporary = output.with_suffix(".tmp" + output.suffix)

    try:
        with temporary.open("w", newline="", encoding="utf-8") as stream:
            writer = csv.DictWriter(stream, fieldnames=OUTPUT_COLUMNS)
            writer.writeheader()
            writer.writerows(rows)

        temporary.replace(output)
    finally:
        if temporary.exists():
            temporary.unlink()


def actual_generation_output_path(output_dir: Path) -> Path:
    """Return the path of the actual-generation CSV file."""

    return output_dir / ACTUAL_GENERATION_FILENAME


def installed_renewables_output_path(output_dir: Path, year: int) -> Path:
    """Return the path of an installed-renewables CSV file."""

    return output_dir / INSTALLED_RENEWABLES_FILENAME.format(year=year)


def installed_renewable_export_params(year: int) -> dict[str, str]:
    """Create the information needed to request one annual file."""

    return {
        "f": "xlsx",
        "filterDataset": INSTALLED_RENEWABLE_DATASET,
        "filterYear": str(year),
        "filterFinalTrendTypes": INSTALLED_RENEWABLE_TREND,
        "orderByColumn": "Anno",
        "orderByDir": "desc",
        "pageSize": str(MAX_DOWNLOAD_CENTER_ROWS),
        "db": "dati"
    }


def selected_installed_renewable_years(
    requested_start: int,
    requested_end: int
) -> list[int]:
    """Choose the requested years that are available from Terna."""

    if requested_end < requested_start:
        raise ValueError("The end year must not be earlier than the start year.")

    first_year = max(
        requested_start,
        FIRST_AVAILABLE_RENEWABLE_YEAR
    )
    years = list(range(first_year, requested_end + 1))

    if not years:
        print(
            "WARNING: no installed-renewables data are available for the "
            "requested years.",
            file=sys.stderr
        )

        return []

    if first_year != requested_start:
        print(
            "WARNING: installed-renewables data are unavailable before "
            f"{FIRST_AVAILABLE_RENEWABLE_YEAR}; starting from "
            f"{FIRST_AVAILABLE_RENEWABLE_YEAR}.",
            file=sys.stderr
        )

    return years


def remove_empty_final_cells(row: tuple[object, ...]) -> tuple[object, ...]:
    """Remove empty cells added after the actual Excel data."""

    cells = list(row)

    while cells and cells[-1] is None:
        cells.pop()

    return tuple(cells)


def read_installed_renewable_workbook(
    path: Path,
    year: int
) -> list[tuple[str, ...]]:
    """Read and check one annual Terna Excel file."""

    workbook = load_workbook(
        path,
        read_only=True,
        data_only=True
    )

    try:
        rows = workbook.active.iter_rows(values_only=True)
        header = remove_empty_final_cells(tuple(next(rows, ())))

        if header != INSTALLED_RENEWABLE_SOURCE_HEADERS:
            raise RuntimeError(
                "Unexpected XLSX header from Terna: "
                f"expected {INSTALLED_RENEWABLE_SOURCE_HEADERS!r}, "
                f"received {header!r}."
            )

        data_rows: list[tuple[str, ...]] = []

        for workbook_row in rows:
            row = remove_empty_final_cells(tuple(workbook_row))

            if not row or str(row[0]).strip() != str(year):
                continue

            if len(row) != len(INSTALLED_RENEWABLE_SOURCE_HEADERS):
                raise RuntimeError(
                    f"Terna workbook row has {len(row)} cells; "
                    f"expected {len(INSTALLED_RENEWABLE_SOURCE_HEADERS)}."
                )

            data_rows.append(
                tuple(
                    "" if value is None else str(value)
                    for value in row
                )
            )
    finally:
        workbook.close()

    if not data_rows:
        raise RuntimeError(f"Terna workbook for {year} contains no data rows.")

    return data_rows


def write_installed_renewable_csv_atomic(
    rows: list[tuple[str, ...]],
    output: Path
) -> None:

    """Save installed-renewables rows to a CSV file."""
    temporary = output.with_suffix(".tmp" + output.suffix)

    try:
        with temporary.open("w", newline="", encoding="utf-8") as stream:
            writer = csv.writer(stream)
            writer.writerow(INSTALLED_RENEWABLE_OUTPUT_HEADERS)
            writer.writerows(rows)
        temporary.replace(output)
    finally:
        if temporary.exists():
            temporary.unlink()


def download_installed_renewable_csv(
    session: requests.Session,
    timeout: float,
    year: int,
    output: Path
) -> int:
    """Download one annual Excel file and save it as CSV."""

    output.parent.mkdir(parents=True, exist_ok=True)
    temporary = output.with_suffix(".tmp.xlsx")

    try:
        with session.get(
            DOWNLOAD_CENTER_RECORDS_URL,
            params=installed_renewable_export_params(year),
            stream=True,
            timeout=timeout
        ) as response:
            if not response.ok:
                raise RuntimeError(
                    "Terna XLSX export failed with HTTP "
                    f"{response.status_code}: {response.text[:800]}."
                )

            with temporary.open("wb") as stream:
                for chunk in response.iter_content(chunk_size=1024 * 1024):
                    if chunk:
                        stream.write(chunk)

        rows = read_installed_renewable_workbook(temporary, year)
        write_installed_renewable_csv_atomic(rows, output)

        return len(rows)
    finally:
        if temporary.exists():
            temporary.unlink()


def download_actual_generation(
    start: date,
    end: date,
    output_dir: Path,
    timeout: float,
    request_interval: float
) -> None:
    """Download, check, and save actual-generation data."""

    if request_interval < 0.0:
        raise ValueError("--request-interval cannot be negative.")

    client_id = os.environ.get(CLIENT_ID_ENV, "").strip()
    client_secret = os.environ.get(CLIENT_SECRET_ENV, "").strip()

    if not client_id or not client_secret:
        raise RuntimeError(
            f"Set {CLIENT_ID_ENV} and {CLIENT_SECRET_ENV} before running this "
            "script. Do not store Terna credentials in the repository."
        )

    client = TernaClient(
        client_id,
        client_secret,
        timeout,
        request_interval
    )
    normalized: list[dict[str, str]] = []
    chunks = create_month_list(start, end)

    print("Terna Actual Generation API.")
    print(f"Source: {SOURCE}.")
    print(f"Period: {start} -> {end} (inclusive local dates).")

    for index, (chunk_start, chunk_end) in enumerate(chunks, start=1):
        print(
            f"[{index:02d}/{len(chunks):02d}] "
            f"{chunk_start} -> {chunk_end}.",
            flush=True
        )
        raw_rows = client.actual_generation(chunk_start, chunk_end)
        normalized.extend(normalize_item(item) for item in raw_rows)

    rows = validate_and_sort(normalized)
    output = actual_generation_output_path(output_dir.resolve())
    write_csv_atomic(rows, output)

    print(f"Saved: {output}.")
    print(f"Rows: {len(rows):,}.")
    print(f"UTC range: {rows[0]['time_utc']} -> {rows[-1]['time_utc']}.")


def download_installed_renewables(
    start_year: int,
    end_year: int,
    output_dir: Path,
    timeout: float
) -> None:
    """Download each annual export between two years."""

    years = selected_installed_renewable_years(start_year, end_year)

    if not years:
        return

    session = build_session(download_center=True)
    output_dir = output_dir.resolve()

    print("Terna Download Center: Installed Renewables.")
    print(f"View: {INSTALLED_RENEWABLE_TREND}.")
    print(f"First available year: {FIRST_AVAILABLE_RENEWABLE_YEAR}.")
    print(
        "Selected years: " + ", ".join(str(year) for year in years) + "."
    )
    print(
        "Each annual file contains monthly changes, not cumulative installed "
        "capacity."
    )

    for index, year in enumerate(years, start=1):
        output = installed_renewables_output_path(output_dir, year)
        print(f"[{index:02d}/{len(years):02d}] {year}.", flush=True)
        rows = download_installed_renewable_csv(session, timeout, year, output)
        print(f"Saved: {output} ({rows:,} data rows).")


def main() -> int:
    """Start the Terna download chosen by the user."""

    args = parse_args()

    if args.timeout <= 0.0:
        raise ValueError("--timeout must be positive.")

    if args.command == "actual-generation":
        start, end = validate_dates(args.start, args.end)
        download_actual_generation(
            start,
            end,
            args.output_dir,
            args.timeout,
            args.request_interval
        )
    elif args.command == "installed-renewables":
        start, end = validate_dates(args.start, args.end)
        download_installed_renewables(
            start.year,
            end.year,
            args.output_dir,
            args.timeout
        )
    elif args.command == "all":
        start, end = validate_dates(args.start, args.end)
        download_actual_generation(
            start,
            end,
            args.output_dir,
            args.timeout,
            args.request_interval
        )
        print()
        download_installed_renewables(
            start.year,
            end.year,
            args.output_dir,
            args.timeout
        )
    else:
        raise AssertionError(f"Unhandled command: {args.command}.")

    return 0


if __name__ == "__main__":
    try:
        raise SystemExit(main())
    except (OSError, ValueError, RuntimeError, requests.RequestException) as exc:
        print(f"ERROR: {exc}", file=sys.stderr)
        raise SystemExit(1) from exc
